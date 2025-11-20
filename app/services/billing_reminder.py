from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, List, Sequence

from openpyxl import load_workbook

from app.models.reminder import (
    BillingReminderRequest,
    BillingReminderResponse,
    ReminderDispatchResult,
    ReminderStatus,
)
from app.services.email_client import EmailClient, EmailClientError
from app.services.email_templates import (
    get_billing_reminder_html,
    get_billing_reminder_text,
)
from app.services.waha_client import WahaClient, WahaClientError


class BillingReminderError(Exception):
    """Base exception for reminder operations."""

    pass


class BillingSheetError(BillingReminderError):
    """Raised when the XLSX file cannot be processed."""

    pass


class BillingRowError(BillingReminderError):
    """Raised when a specific row has invalid data and should be skipped."""

    pass


@dataclass
class BillingRecord:
    client_name: str
    whatsapp_number: str
    email: str | None
    due_date: date


class BillingReminderService:
    """Read the billing XLSX and dispatch WhatsApp and email reminders."""

    EXCEL_EPOCH = datetime(1899, 12, 30)

    CLIENT_HEADERS = {"cliente", "client", "clientenome", "nome", "name"}
    PHONE_HEADERS = {"telefone", "phone", "whatsapp", "numerowhatsapp", "numero"}
    EMAIL_HEADERS = {"email", "e-mail", "mail", "correio"}
    DUE_HEADERS = {"vencimento", "datavencimento", "data", "duedate"}

    DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y")

    def __init__(
        self,
        default_sheet_path: str,
        reminder_days: Sequence[int],
        waha_client: WahaClient,
        email_client: EmailClient | None = None,
        email_enabled: bool = False,
    ) -> None:
        if not reminder_days:
            raise ValueError("reminder_days cannot be empty")

        self._default_sheet_path = default_sheet_path
        self._reminder_days = sorted(set(reminder_days))
        self._waha_client = waha_client
        self._email_client = email_client
        self._email_enabled = email_enabled

    def run(self, request: BillingReminderRequest) -> BillingReminderResponse:
        sheet_path = Path(request.sheet_path or self._default_sheet_path).expanduser()
        reference_date = request.reference_date or date.today()

        records = self._load_records(sheet_path)
        total_rows = len(records)

        results: List[ReminderDispatchResult] = []
        eligible_rows = 0
        dispatched = 0

        for record in sorted(records, key=lambda rec: rec.due_date):
            days_until_due = (record.due_date - reference_date).days
            if days_until_due not in self._reminder_days:
                continue
            if days_until_due < 0:
                continue

            eligible_rows += 1
            message = self._build_message(record, days_until_due)

            # Send WhatsApp
            whatsapp_status = ReminderStatus.SKIPPED
            whatsapp_detail = None
            if request.dry_run:
                whatsapp_status = ReminderStatus.DRY_RUN
                whatsapp_detail = "Dry-run: WhatsApp não enviado."
            else:
                try:
                    api_result = self._waha_client.send_text_message(
                        recipient=record.whatsapp_number,
                        message=message,
                        sender=request.sender_whatsapp_number,
                    )
                    whatsapp_status = ReminderStatus.SENT
                    whatsapp_detail = api_result.get("message", "Mensagem registrada no WAHA.")
                except WahaClientError as exc:
                    whatsapp_status = ReminderStatus.FAILED
                    whatsapp_detail = str(exc)

            # Send Email
            email_status = ReminderStatus.SKIPPED
            email_detail = None
            if self._email_enabled and self._email_client and record.email:
                if request.dry_run:
                    email_status = ReminderStatus.DRY_RUN
                    email_detail = "Dry-run: Email não enviado."
                else:
                    try:
                        html_content = get_billing_reminder_html(
                            record.client_name, record.due_date, days_until_due
                        )
                        text_content = get_billing_reminder_text(
                            record.client_name, record.due_date, days_until_due
                        )
                        email_result = self._email_client.send_email(
                            to_email=record.email,
                            subject=f"Lembrete de Boleto - Vence em {days_until_due} dia(s)",
                            html_content=html_content,
                            text_content=text_content,
                        )
                        email_status = ReminderStatus.SENT
                        email_detail = email_result.get("message", "Email enviado com sucesso.")
                    except EmailClientError as exc:
                        email_status = ReminderStatus.FAILED
                        email_detail = str(exc)
            elif record.email is None:
                email_detail = "Email não informado na planilha."

            # Determine overall status (SENT if at least one succeeded)
            overall_status = whatsapp_status
            if email_status == ReminderStatus.SENT and whatsapp_status != ReminderStatus.SENT:
                overall_status = ReminderStatus.SENT
            elif whatsapp_status == ReminderStatus.SENT:
                overall_status = ReminderStatus.SENT

            # Build detail message
            detail_parts = []
            if whatsapp_detail:
                detail_parts.append(f"WhatsApp: {whatsapp_detail}")
            if email_detail:
                detail_parts.append(f"Email: {email_detail}")
            detail = " | ".join(detail_parts) if detail_parts else None

            if overall_status in {ReminderStatus.SENT, ReminderStatus.DRY_RUN}:
                dispatched += 1

            results.append(
                ReminderDispatchResult(
                    client_name=record.client_name,
                    whatsapp_number=record.whatsapp_number,
                    due_date=record.due_date,
                    days_until_due=days_until_due,
                    status=overall_status,
                    message_preview=message[:120],
                    detail=detail,
                )
            )

        return BillingReminderResponse(
            sheet_path=str(sheet_path),
            reference_date=reference_date,
            days_watched=self._reminder_days,
            dry_run=request.dry_run,
            total_rows=total_rows,
            eligible_rows=eligible_rows,
            dispatched=dispatched,
            results=results,
        )

    def _load_records(self, sheet_path: Path) -> List[BillingRecord]:
        if not sheet_path.exists():
            raise BillingSheetError(f"Planilha nao encontrada em {sheet_path}")

        try:
            workbook = load_workbook(sheet_path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl specific errors
            raise BillingSheetError(f"Falha ao abrir {sheet_path}: {exc}") from exc

        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            return []

        header = rows[0]
        indexes = self._resolve_indexes(header)

        records: List[BillingRecord] = []
        for row in rows[1:]:
            try:
                record = self._build_record(row, indexes)
            except BillingRowError:
                continue
            records.append(record)

        return records

    def _resolve_indexes(self, header_row: Iterable) -> dict[str, int]:
        indexes: dict[str, int] = {}
        for idx, raw_name in enumerate(header_row):
            normalized = self._normalize_header(raw_name)
            if not normalized:
                continue
            if normalized in self.CLIENT_HEADERS:
                indexes["client_name"] = idx
            elif normalized in self.PHONE_HEADERS:
                indexes["whatsapp_number"] = idx
            elif normalized in self.EMAIL_HEADERS:
                indexes["email"] = idx
            elif normalized in self.DUE_HEADERS:
                indexes["due_date"] = idx

        missing = {"client_name", "whatsapp_number", "due_date"} - indexes.keys()
        if missing:
            raise BillingSheetError(
                "Colunas obrigatorias ausentes no cabecalho: "
                + ", ".join(sorted(missing))
            )
        return indexes

    def _build_record(self, row: Iterable, indexes: dict[str, int]) -> BillingRecord:
        client_value = row[indexes["client_name"]]
        phone_value = row[indexes["whatsapp_number"]]
        due_value = row[indexes["due_date"]]

        if not client_value or not phone_value or not due_value:
            raise BillingRowError("Linha com valores obrigatorios vazios.")

        client_name = str(client_value).strip()
        whatsapp_number = self._sanitize_phone(str(phone_value))
        due_date = self._parse_due_date(due_value)

        if not client_name or not whatsapp_number:
            raise BillingRowError("Linha com cliente ou telefone invalido.")

        # Email is optional
        email = None
        if "email" in indexes:
            email_value = row[indexes["email"]]
            if email_value:
                email = str(email_value).strip().lower()
                if not email or "@" not in email:
                    email = None

        return BillingRecord(
            client_name=client_name,
            whatsapp_number=whatsapp_number,
            email=email,
            due_date=due_date,
        )

    def _parse_due_date(self, value) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            excel_date = self.EXCEL_EPOCH + timedelta(days=float(value))
            return excel_date.date()
        if isinstance(value, str):
            cleaned = value.strip()
            for fmt in self.DATE_FORMATS:
                try:
                    return datetime.strptime(cleaned, fmt).date()
                except ValueError:
                    continue
            try:
                as_float = float(cleaned)
            except ValueError:
                pass
            else:
                excel_date = self.EXCEL_EPOCH + timedelta(days=as_float)
                return excel_date.date()

        raise BillingRowError(f"Nao foi possivel converter a data: {value!r}")

    def _build_message(self, record: BillingRecord, days_until_due: int) -> str:
        if days_until_due <= 1:
            return (
                f"Ola {record.client_name}, o seu boleto vence em "
                f"{record.due_date:%d/%m/%Y}. Falta 1 dia para o vencimento."
            )
        return (
            f"Ola {record.client_name}, faltam {days_until_due} dias para o "
            f"vencimento do seu boleto ({record.due_date:%d/%m/%Y})."
        )

    @staticmethod
    def _normalize_header(value) -> str:
        if value is None:
            return ""
        return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())

    @staticmethod
    def _sanitize_phone(phone: str) -> str:
        digits = "".join(ch for ch in phone if ch.isdigit())
        if phone.strip().startswith("+") and digits:
            return f"+{digits}"
        return digits
